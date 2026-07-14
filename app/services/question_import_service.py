from __future__ import annotations

import re
from pathlib import Path
from xml.etree import ElementTree as ET
from zipfile import ZipFile

try:
    from docx import Document as DocxDocument
except Exception:
    DocxDocument = None

try:
    from openpyxl import load_workbook as openpyxl_load_workbook
except Exception:
    openpyxl_load_workbook = None


SECTION_TRUE_FALSE = "\u041f\u0420\u0410\u0412\u0414\u0410 \u0418\u041b\u0418 \u041f\u0418\u0417\u0414\u0401\u0416?"
SECTION_SINGLE_CHOICE = "\u0412\u0410\u0420\u0418\u0410\u041d\u0422\u042b \u041e\u0422\u0412\u0415\u0422\u041e\u0412"
SECTION_FREE_TEXT = "\u0412\u041e\u041f\u0420\u041e\u0421\u042b \u0411\u0415\u0417 \u0412\u0410\u0420\u0418\u0410\u041d\u0422\u041e\u0412"
SECTION_NOTE = "\u041f\u0420\u0418\u041c\u0415\u0427\u0410\u041d\u0418\u0415"

SECTION_TYPE_MAP = {
    SECTION_TRUE_FALSE: "true_false",
    SECTION_SINGLE_CHOICE: "single_choice",
    SECTION_FREE_TEXT: "free_text",
}

MEDIA_TYPE_MAP = {
    "\u0444\u043e\u0442\u043e": "image",
    "image": "image",
    "\u043a\u0430\u0440\u0442\u0438\u043d\u043a\u0430": "image",
    "audio": "audio",
    "\u0430\u0443\u0434\u0438\u043e": "audio",
    "\u0437\u0432\u0443\u043a": "audio",
    "video": "video",
    "\u0432\u0438\u0434\u0435\u043e": "video",
    "none": "none",
    "\u043f\u0443\u0441\u0442\u043e": "none",
    "": "none",
}

WORD_NS = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
SHEET_NS = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "pkg": "http://schemas.openxmlformats.org/package/2006/relationships",
}

INLINE_MEDIA_PATTERNS = [
    re.compile(
        r'\((?P<label>\u0424\u043e\u0442\u043e|\u0444\u043e\u0442\u043e|\u0412\u0438\u0434\u0435\u043e|\u0432\u0438\u0434\u0435\u043e)\s*[\u00ab"](?P<ref>[^\u00bb"]+)[\u00bb"]\)',
    ),
    re.compile(
        r'\((?P<label>\u0424\u043e\u0442\u043e|\u0444\u043e\u0442\u043e|\u0412\u0438\u0434\u0435\u043e|\u0432\u0438\u0434\u0435\u043e)\s+(?P<ref>[^)]+)\)',
    ),
    re.compile(
        r'(?P<label>\u0424\u043e\u0442\u043e|\u0444\u043e\u0442\u043e|\u0412\u0438\u0434\u0435\u043e|\u0432\u0438\u0434\u0435\u043e)\s*[\u00ab"](?P<ref>[^\u00bb"]+)[\u00bb"]',
    ),
]


def _normalize_whitespace(value) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").strip()
    return re.sub(r"\s+", " ", text)


def _normalize_section_title(value: str) -> str:
    text = _normalize_whitespace(value)
    if not text:
        return ""
    text = re.sub(r"^#+\s*", "", text)
    comparable = text.replace("\u0401", "\u0415").upper().rstrip(":")
    if "\u0412\u0410\u0420\u0418\u0410\u041d\u0422\u042b \u041e\u0422\u0412\u0415\u0422\u041e\u0412" in comparable:
        return SECTION_SINGLE_CHOICE
    if "\u0412\u041e\u041f\u0420\u041e\u0421\u042b \u0411\u0415\u0417 \u0412\u0410\u0420\u0418\u0410\u041d\u0422\u041e\u0412" in comparable:
        return SECTION_FREE_TEXT
    if "\u041f\u0420\u0410\u0412\u0414\u0410 \u0418\u041b\u0418 \u041f\u0418\u0417\u0414" in comparable:
        return SECTION_TRUE_FALSE
    if "\u041f\u0420\u0418\u041c\u0415\u0427\u0410\u041d\u0418\u0415" in comparable:
        return SECTION_NOTE
    return text.rstrip(":")


def _infer_media_type(media_type: str, media_ref: str) -> str:
    media_type = _normalize_whitespace(media_type).lower()
    media_ref = _normalize_whitespace(media_ref)
    if media_type in MEDIA_TYPE_MAP:
        return MEDIA_TYPE_MAP[media_type]

    lowered_ref = media_ref.lower()
    if "\u0432\u0438\u0434\u0435\u043e" in lowered_ref or lowered_ref.endswith((".mp4", ".mov", ".avi", ".mkv")):
        return "video"
    if "\u0430\u0443\u0434\u0438\u043e" in lowered_ref or lowered_ref.endswith((".mp3", ".wav", ".ogg", ".m4a")):
        return "audio"
    if "\u0444\u043e\u0442\u043e" in lowered_ref or lowered_ref.endswith((".jpg", ".jpeg", ".png", ".webp")):
        return "image"
    return "image" if media_ref else "none"


def _extract_inline_media(*values: str) -> tuple[str, str]:
    for value in values:
        text = _normalize_whitespace(value)
        if not text:
            continue
        for pattern in INLINE_MEDIA_PATTERNS:
            match = pattern.search(text)
            if not match:
                continue
            label = _normalize_whitespace(match.group("label")).lower()
            ref = _normalize_whitespace(match.group("ref")).strip(" .")
            if ref:
                return _infer_media_type(label, ref), ref
    return "none", ""


def _append_text(base: str, extra: str) -> str:
    base = _normalize_whitespace(base)
    extra = _normalize_whitespace(extra)
    if not extra:
        return base
    if not base:
        return extra
    return f"{base} {extra}"


def _build_question_code(question_type: str, index: int) -> str:
    return f"import_{question_type}_{index:03d}"


def _read_docx_paragraphs_from_xml(file_path) -> list[tuple[str, bool]]:
    paragraphs: list[tuple[str, bool]] = []
    with ZipFile(file_path, "r") as archive:
        xml_bytes = archive.read("word/document.xml")
    root = ET.fromstring(xml_bytes)
    for paragraph in root.findall(".//w:p", WORD_NS):
        text_parts: list[str] = []
        has_italic = False
        for run in paragraph.findall("./w:r", WORD_NS):
            texts = [node.text or "" for node in run.findall(".//w:t", WORD_NS)]
            if texts:
                text_parts.append("".join(texts))
            if run.find("./w:rPr/w:i", WORD_NS) is not None:
                has_italic = True
        text = _normalize_whitespace("".join(text_parts))
        if text:
            paragraphs.append((text, has_italic))
    return paragraphs


def _extract_shared_strings(archive: ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in archive.namelist():
        return []
    root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    values = []
    for item in root.findall(".//main:si", SHEET_NS):
        text = "".join(node.text or "" for node in item.findall(".//main:t", SHEET_NS))
        values.append(text)
    return values


def _resolve_questions_sheet_path(archive: ZipFile) -> str:
    workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
    rel_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    archive_paths = set(archive.namelist())
    rel_map = {
        rel.attrib.get("Id"): rel.attrib.get("Target")
        for rel in rel_root.findall(".//pkg:Relationship", SHEET_NS)
    }
    for sheet in workbook_root.findall(".//main:sheet", SHEET_NS):
        if sheet.attrib.get("name") == "questions":
            rel_id = sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
            target = rel_map.get(rel_id)
            if target:
                normalized_target = target.replace("\\", "/").lstrip("/")
                candidates = []
                if normalized_target.startswith("xl/"):
                    candidates.append(normalized_target)
                else:
                    candidates.append(f"xl/{normalized_target}")
                candidates.append(normalized_target)

                unique_candidates = list(dict.fromkeys(candidates))
                for candidate in unique_candidates:
                    if candidate in archive_paths:
                        return candidate

                raise ValueError(
                    f'\u041f\u0443\u0442\u044c \u043b\u0438\u0441\u0442\u0430 "questions" \u043d\u0435 \u043d\u0430\u0439\u0434\u0435\u043d \u0432 XLSX: '
                    f"target={target!r}, candidates={unique_candidates!r}"
                )
    raise ValueError('\u041b\u0438\u0441\u0442 "questions" \u043d\u0435 \u043d\u0430\u0439\u0434\u0435\u043d')


def _xlsx_column_index(cell_ref: str) -> int:
    match = re.match(r"^([A-Z]+)", cell_ref or "")
    if not match:
        return 0
    name = match.group(1)
    value = 0
    for char in name:
        value = value * 26 + (ord(char) - 64)
    return value - 1


def _read_xlsx_rows_from_xml(file_path) -> list[list[str]]:
    with ZipFile(file_path, "r") as archive:
        shared_strings = _extract_shared_strings(archive)
        sheet_path = _resolve_questions_sheet_path(archive)
        root = ET.fromstring(archive.read(sheet_path))

    rows: list[list[str]] = []
    for row in root.findall(".//main:sheetData/main:row", SHEET_NS):
        values: list[str] = []
        for cell in row.findall("./main:c", SHEET_NS):
            column_index = _xlsx_column_index(cell.attrib.get("r", ""))
            while len(values) <= column_index:
                values.append("")
            cell_type = cell.attrib.get("t")
            value = ""
            if cell_type == "inlineStr":
                value = "".join(node.text or "" for node in cell.findall(".//main:is/main:t", SHEET_NS))
            elif cell_type == "s":
                value_node = cell.find("./main:v", SHEET_NS)
                if value_node is not None and value_node.text and value_node.text.isdigit():
                    idx = int(value_node.text)
                    value = shared_strings[idx] if idx < len(shared_strings) else ""
            else:
                value_node = cell.find("./main:v", SHEET_NS)
                value = value_node.text if value_node is not None and value_node.text is not None else ""
            values[column_index] = value
        rows.append(values)
    return rows


def normalize_imported_question(raw_question: dict) -> dict:
    section = _normalize_section_title(raw_question.get("section"))
    prompt = _normalize_whitespace(raw_question.get("prompt") or raw_question.get("question"))
    question_type = _normalize_whitespace(raw_question.get("type") or raw_question.get("question_type")).lower()
    if not question_type:
        question_type = SECTION_TYPE_MAP.get(section, "")

    option_candidates = raw_question.get("options")
    if not isinstance(option_candidates, list):
        option_candidates = [
            raw_question.get("option_a"),
            raw_question.get("option_b"),
            raw_question.get("option_c"),
            raw_question.get("option_d"),
        ]
    options = [_normalize_whitespace(item) for item in option_candidates if _normalize_whitespace(item)]
    if question_type == "free_text":
        options = []

    correct_answer = _normalize_whitespace(raw_question.get("correct_answer"))
    explanation = _normalize_whitespace(raw_question.get("explanation"))
    media_ref = _normalize_whitespace(raw_question.get("media_ref"))
    raw_media_type = _normalize_whitespace(raw_question.get("media_type"))

    inline_media_type, inline_media_ref = _extract_inline_media(prompt, explanation, media_ref)
    if inline_media_ref and not media_ref:
        media_ref = inline_media_ref
    if not raw_media_type and inline_media_type != "none":
        raw_media_type = inline_media_type

    media_type = _infer_media_type(raw_media_type, media_ref)
    if media_type == "none" and media_ref:
        media_type = "image"
    is_media_question = bool(media_ref) or media_type != "none"

    errors: list[str] = []
    if not prompt:
        errors.append("\u041d\u0435 \u0437\u0430\u043f\u043e\u043b\u043d\u0435\u043d \u0442\u0435\u043a\u0441\u0442 \u0432\u043e\u043f\u0440\u043e\u0441\u0430")
    if not correct_answer:
        errors.append("\u041d\u0435 \u0443\u043a\u0430\u0437\u0430\u043d \u043f\u0440\u0430\u0432\u0438\u043b\u044c\u043d\u044b\u0439 \u043e\u0442\u0432\u0435\u0442")
    if question_type == "single_choice" and not options:
        errors.append("single_choice \u0432\u043e\u043f\u0440\u043e\u0441 \u0434\u043e\u043b\u0436\u0435\u043d \u0441\u043e\u0434\u0435\u0440\u0436\u0430\u0442\u044c \u0432\u0430\u0440\u0438\u0430\u043d\u0442\u044b \u043e\u0442\u0432\u0435\u0442\u043e\u0432")

    return {
        "question_code": raw_question.get("question_code") or _build_question_code(question_type or "question", int(raw_question.get("index") or 0)),
        "section": section,
        "type": question_type,
        "prompt": prompt,
        "options": options,
        "correct_answer": correct_answer,
        "explanation": explanation,
        "media_type": media_type,
        "media_ref": media_ref,
        "is_media_question": is_media_question,
        "difficulty": _normalize_whitespace(raw_question.get("difficulty")),
        "role_code": _normalize_whitespace(raw_question.get("role_code")),
        "round_code": _normalize_whitespace(raw_question.get("round_code")),
        "tags": _normalize_whitespace(raw_question.get("tags")),
        "has_errors": bool(errors),
        "errors": errors,
    }


def parse_questions_docx(file_path) -> list[dict]:
    if DocxDocument is not None:
        paragraph_rows = [
            (
                _normalize_whitespace(paragraph.text),
                any(bool(run.italic) for run in paragraph.runs if _normalize_whitespace(run.text)),
            )
            for paragraph in DocxDocument(str(file_path)).paragraphs
        ]
    else:
        paragraph_rows = _read_docx_paragraphs_from_xml(file_path)

    results: list[dict] = []
    current_section = ""
    current_question: dict | None = None

    def flush_current():
        nonlocal current_question
        if current_question:
            current_question["index"] = len(results) + 1
            results.append(normalize_imported_question(current_question))
            current_question = None

    for raw_line, has_italic in paragraph_rows:
        line = _normalize_whitespace(raw_line)
        if not line:
            continue

        heading_line = _normalize_section_title(line)
        if heading_line in SECTION_TYPE_MAP:
            flush_current()
            current_section = heading_line
            continue

        if heading_line == SECTION_NOTE:
            flush_current()
            current_section = ""
            continue

        question_match = re.match(r"^(\d+)\.\s+(.+)$", line)
        if question_match:
            flush_current()
            current_question = {
                "index": int(question_match.group(1)),
                "section": current_section,
                "question": question_match.group(2).strip(),
                "options": [],
                "correct_answer": "",
                "explanation": "",
                "media_type": "",
                "media_ref": "",
            }
            continue

        if not current_question:
            continue

        question_type = SECTION_TYPE_MAP.get(current_section, "")

        if line.lower().startswith("\u043f\u0440\u0430\u0432\u0438\u043b\u044c\u043d\u044b\u0439 \u043e\u0442\u0432\u0435\u0442:"):
            current_question["correct_answer"] = _normalize_whitespace(line.split(":", 1)[1])
            continue

        if line.lower().startswith("\u043f\u043e\u044f\u0441\u043d\u0435\u043d\u0438\u0435:"):
            current_question["explanation"] = _append_text(current_question.get("explanation", ""), line.split(":", 1)[1])
            continue

        if line.lower().startswith("\u043c\u0435\u0434\u0438\u0430:"):
            media_value = _normalize_whitespace(line.split(":", 1)[1])
            current_question["media_ref"] = "" if media_value.lower() == "\u043f\u0443\u0441\u0442\u043e" else media_value
            current_question["media_type"] = _infer_media_type("", current_question["media_ref"])
            continue

        option_match = re.match(r"^-\s*(.+)$", line)
        if option_match:
            option_text = _normalize_whitespace(option_match.group(1))
            if question_type == "free_text":
                if not _normalize_whitespace(current_question.get("correct_answer")):
                    current_question["correct_answer"] = option_text
                else:
                    current_question["explanation"] = _append_text(current_question.get("explanation", ""), option_text)
            else:
                current_question.setdefault("options", []).append(option_text)
                if has_italic and not _normalize_whitespace(current_question.get("correct_answer")):
                    current_question["correct_answer"] = option_text
            continue

        if question_type == "free_text" and has_italic and not _normalize_whitespace(current_question.get("correct_answer")):
            current_question["correct_answer"] = line
            continue

        if line.startswith("(") and line.endswith(")"):
            current_question["explanation"] = _append_text(current_question.get("explanation", ""), line)
            continue

        if question_type == "free_text" and not _normalize_whitespace(current_question.get("correct_answer")):
            current_question["question"] = _append_text(current_question.get("question", ""), line)
            continue

        current_question["explanation"] = _append_text(current_question.get("explanation", ""), line)

    flush_current()
    return results


def parse_questions_xlsx(file_path) -> list[dict]:
    if openpyxl_load_workbook is not None:
        workbook = openpyxl_load_workbook(filename=str(file_path), data_only=True)
        if "questions" not in workbook.sheetnames:
            raise ValueError('\u041b\u0438\u0441\u0442 "questions" \u043d\u0435 \u043d\u0430\u0439\u0434\u0435\u043d')
        sheet = workbook["questions"]
        rows = list(sheet.iter_rows(values_only=True))
    else:
        rows = _read_xlsx_rows_from_xml(file_path)

    if not rows:
        return []

    headers = [_normalize_whitespace(value) for value in rows[0]]
    header_map = {header: idx for idx, header in enumerate(headers) if header}

    results: list[dict] = []
    for row_index, row in enumerate(rows[1:], start=1):
        raw = {header: row[idx] if idx < len(row) else None for header, idx in header_map.items()}
        if not any(_normalize_whitespace(value) for value in raw.values()):
            continue

        raw_question = {
            "index": row_index,
            "section": raw.get("section"),
            "question_type": raw.get("question_type"),
            "question": raw.get("question"),
            "option_a": raw.get("option_a"),
            "option_b": raw.get("option_b"),
            "option_c": raw.get("option_c"),
            "option_d": raw.get("option_d"),
            "correct_answer": raw.get("correct_answer"),
            "explanation": raw.get("explanation"),
            "media_type": raw.get("media_type"),
            "media_ref": raw.get("media_ref"),
            "difficulty": raw.get("difficulty"),
            "role_code": raw.get("role_code"),
            "round_code": raw.get("round_code"),
            "tags": raw.get("tags"),
        }
        results.append(normalize_imported_question(raw_question))

    return results


def build_questions_import_preview(file_path) -> dict:
    path = Path(file_path)
    suffix = path.suffix.lower()
    if suffix == ".docx":
        questions = parse_questions_docx(path)
        source_type = "docx"
    elif suffix == ".xlsx":
        questions = parse_questions_xlsx(path)
        source_type = "xlsx"
    else:
        raise ValueError("\u041f\u043e\u0434\u0434\u0435\u0440\u0436\u0438\u0432\u0430\u044e\u0442\u0441\u044f \u0442\u043e\u043b\u044c\u043a\u043e \u0444\u0430\u0439\u043b\u044b DOCX \u0438 XLSX")

    sections = {
        "true_false": 0,
        "single_choice": 0,
        "free_text": 0,
    }
    for item in questions:
        item_type = item.get("type")
        if item_type in sections:
            sections[item_type] += 1

    return {
        "ok": True,
        "source_type": source_type,
        "questions_count": len(questions),
        "sections": sections,
        "questions": questions,
    }


def select_questions_by_limits(
    questions: list[dict],
    *,
    true_false_limit: int = 5,
    single_choice_limit: int = 5,
    free_text_limit: int = 3,
    media_limit: int = 0,
    prefer_media: bool = False,
) -> dict:
    limits = {
        "true_false": max(0, int(true_false_limit or 0)),
        "single_choice": max(0, int(single_choice_limit or 0)),
        "free_text": max(0, int(free_text_limit or 0)),
    }
    selected: list[dict] = []
    selected_codes: set[str] = set()
    by_type = {
        "true_false": 0,
        "single_choice": 0,
        "free_text": 0,
    }

    for question_type in ("true_false", "single_choice", "free_text"):
        type_items = [item for item in questions if item.get("type") == question_type]
        if prefer_media:
            type_items = sorted(
                type_items,
                key=lambda item: (not bool(item.get("is_media_question")), questions.index(item)),
            )
        picked = type_items[: limits[question_type]]
        selected.extend(picked)
        selected_codes.update(
            str(item.get("question_code") or "")
            for item in picked
            if str(item.get("question_code") or "")
        )
        by_type[question_type] = len(picked)

    media_limit_value = max(0, int(media_limit or 0))
    media_selected: list[dict] = []
    if media_limit_value > 0:
        media_candidates = [
            item
            for item in questions
            if item.get("is_media_question")
            and str(item.get("question_code") or "") not in selected_codes
        ]
        media_selected = media_candidates[:media_limit_value]
        selected.extend(media_selected)
        selected_codes.update(
            str(item.get("question_code") or "")
            for item in media_selected
            if str(item.get("question_code") or "")
        )

    return {
        "selected_count": len(selected),
        "by_type": by_type,
        "media_count": len(media_selected),
        "selected_questions": selected,
    }
